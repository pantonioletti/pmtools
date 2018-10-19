import openpyxl as xl
from openpyxl.styles import PatternFill, Alignment, Font
from functools import reduce
import datetime as dt
import array as arr
import sys
import os

'''
Resource is identified by name.
It has forecast by date and actuals buy date and rate
'''
class Resource:
    def __init__(self, name):
        self.name = name
        self.forecast = dict()
        self.actuals = dict()

    def add_actual(self, date, actual, rate):
        #if rate not in self.rates:
        #    self.rates.append(rate)
        if date in self.actuals:
            if rate in self.actuals[date]:
                self.actuals[date][rate] += actual
            else:
                self.actuals[date][rate] = actual
        else:
            self.actuals[date] = {rate: actual}

    def add_forecast(self,date, forecast):
        if date in self.forecast:
            self.forecast[date] += forecast
        else:
            self.forecast[date] = forecast

    def toString(self):
        retVal = arr.array('c')
        retVal.fromstring('%s :\n'%(self.name))
        for w in sorted(self.forecast.keys()):
            retVal.fromstring('%d-%d-%d | %d |\n'%(w.year, w.month, w.day,self.forecast[w]))
        return retVal.tostring()

    def getActuals(self):
        return self.actuals

    def getForecast(self):
        return self.forecast

    def getLastRate(self):
        avg = 0.0
        if len(self.actuals) > 0:
            last_date = max(self.actuals.keys())
            revenue = map(lambda x: float(self.actuals[last_date][x])*x, self.actuals[last_date].keys())
            hours = sum((x for x in self.actuals[last_date].values()))
            if hours == 0.0:
                pass
            avg = reduce(lambda x,y: x+y, revenue) / hours
        return round(avg,2)


'''
Returns the Sunday before <date> if <date> is not already Sunday
'''
def getSunday(date):
    ndate = date
    if type(date) is str :
        if len(date) == 10:
            ndate = dt.date(int(date[-4:]), int(date[0:2]), int(date[-7:-5]))
        else:
            ndate = None
    elif type(ndate) is dt.datetime:
        ndate = ndate.date()
    dow = ndate.weekday()
    if dow < 6:
        ndate = ndate - dt.timedelta(dow+1)
    return ndate


'''
This function will process a spreadsheet where columns are:
.__________________________________________________.
|         Column data               | Column index |
.__________________________________________________.
| Client Reporting Unit            	|      0       |
| Client Name	                    |      1       | 
| Project Name	                    |      2       |
| Client PO#	                    |      3       |
| Project Manager	                |      4       |
| Associate Name	                |      5       |
| AssociateId	                    |      6       |
| AssociateType	                    |      7       |
| Period Start Date	                |      8       |
| Entry Date	                    |      9       |
| Task Name	                        |     10       |
| Total Hours	                    |     11       |
| Billable Hours	                |     12       |
| Billing Rule Name	                |     13       |
| Billable Amt. In Project Currency	|     14       |
| Project Currency Name	            |     15       |
| Billable Amt. In USD	            |     16       |
| Timesheet is Submitted	        |     17       |
| Timesheet is Approved           	|     18       |
| Timesheet Workflow State	        |     19       |
| JTRAX Invoiced Status/Ref #	    |     20       |
| Invoice Through Date	            |     21       |
| Service Delivery Location	        |     22       |
| Entry Note	                    |     23       |
| On Hold for Billing	            |     24       |
| On Hold Reason                    |     25       |
.__________________________________________________.
Relevant data to be kept is associate name (column 5), entry date (column 9),  hours (column 11), amount in USD (column 16).
Amount by itself is not kept, it is used to calculate rate (rate= <amount> / <hours>). Hours are consolidated by week/rate.
Weeks start on Sunday.
First row with data is 9. 
'''

def proc_actuals(file, resources, sheet=None, se_dates=[dt.date.today(), dt.date.today()]):
    wb = xl.load_workbook(filename=file,
                          read_only=True,
                          keep_vba=False,
                          data_only=False,
                          guess_types=False,
                          keep_links=True)
    if sheet is None:
        ws = wb.active
    else:
        ws = wb[sheet]
    for row in ws.iter_rows(range_string=None, min_row=9, max_row=None, min_col=None, max_col=None, row_offset=0, column_offset=0):
        if row[9].is_date:
            sunday = getSunday(dt.date(row[9].value.year, row[9].value.month, row[9].value.day))
            if sunday < se_dates[0]:
                se_dates[0] = sunday
            elif sunday > se_dates[1]:
                se_dates[1] = sunday
            name = row[5].value.strip()
            if name not in resources:
                resources[name] = Resource(name)
            hh = float(row[11].value)
            amt = round(float(row[16].value)/hh,2)
            resources[name].add_actual(sunday, hh, amt)


'''
This function will process a spreadsheet where columns are:
.__________________________________.
|   Column data      |Column index |
.__________________________________.
|User Last Name      |     0       |
|User First Name     |     1       |
|Role                |     2       |
|Project             |     3       |
|Date                |     4       |
|Actual Hours        |     5       |
|Total Booking Hours |     6       |
.__________________________________.

Relevant data to be kept is consultant name (columns 0 and 1), date (columns 4) and booking hours (column 6).
Data is consolidated by week. Weeks start on Sunday.
First row with data is 9.
'''
def proc_forecast(file, resources, sheet=None, se_dates = [dt.date.today(), dt.date.today()]):
    wb = xl.load_workbook(filename=file,
                          read_only=True,
                          keep_vba=False,
                          data_only=False,
                          guess_types=False,
                          keep_links=True)

    if sheet is None:
        ws = wb.active
    else:
        ws = wb[sheet]
    for row in ws.iter_rows(range_string=None, min_row=9, max_row=None, min_col=None, max_col=None,
                            row_offset=0, column_offset=0):
        if row[4].is_date:
            date = dt.date(row[4].value.year, row[4].value.month, row[4].value.day)
            sunday = getSunday(date)
            if sunday < se_dates[0]:
                se_dates[0] = sunday
            elif sunday > se_dates[1]:
                se_dates[1] = sunday
            if row[0].value is None:
                name = '{0}'.format(row[2].value.strip())
            else:
                name = '{0}, {1}'.format(row[0].value.strip(), row[1].value.strip())
            if name not in resources:
                resources[name] = Resource(name)
            resources[name].add_forecast(sunday, float(row[6].value))
    wb.close()


def create_date_seq(start_date, end_date):
    dcmap = dict()
    curr_date = start_date
    curr_col = 0
    while curr_date <= end_date:
        dcmap[curr_date] = curr_col
        curr_col += 1
        curr_date = curr_date + dt.timedelta(7)
    curr_col += 1
    return dcmap


def create_headers(ws, title, dcmap):
    ws.title = title
    ws.insert_rows(1, 2)
    f = PatternFill(patternType=None, start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    font = Font(bold=True)
    ws.cell(1,1,'Row Labels')
    ws.cell(1, 2, 'Rate')
    curr_col = 3
    r = Alignment(text_rotation=90, horizontal='center')
    for curr_date in dcmap.keys():
        c = ws.cell(1,dcmap[curr_date]+curr_col,curr_date)
        c.fill = f
        c.alignment = r
        c.font = font
    curr_col += len(dcmap)
    c = ws.cell(1, curr_col, "Total Hours")
    c.fill = f
    r = Alignment(wrap_text=True, horizontal='center')
    c.alignment = r
    c.font = font
    c = ws.cell(1, curr_col+1, "Total Cost")
    c.fill = f
    c.alignment = r
    c.font = font


def setColorToZeros(ws, pf):
    for row in ws.iter_rows(min_row=2, min_col=2):
        if row[0].value == 0.0:
            for col in row:
                col.fill = pf


def addFormulas(ws, shift):
    start_col = 3
    last_col = ws.cell(1, start_col+shift).column
    hours_formula = "=SUM({0}{1}:{2}{1})"
    cost_formula = "={0}{1}*B{1}"
    hours = start_col + shift + 1
    cost = hours + 1
    for r in ws.iter_rows(min_row=2):
        c = ws.cell(r[0].row, hours)
        c.set_explicit_value(hours_formula.format('C',r[0].row,last_col), data_type = 'f')
        h_col = c.column
        c = ws.cell(r[0].row,cost)
        c.set_explicit_value(cost_formula.format(h_col,r[0].row),data_type='f')


def actuals_sheet(ws, resources, date_col):
    row = 2
    start_col = 3
    for res_name in resources.keys():
        ws.cell(row, 1, res_name)
        actuals = res[res_name].getActuals()
        rrmap = dict()
        for date in actuals.keys():
            res_col = date_col[date]+start_col
            res_row = row
            for rate in actuals[date].keys():
                if len(rrmap)==0:
                    rrmap[rate] = res_row
                    ws.cell(res_row, 2, rate)
                else:
                    if rate in rrmap:
                        res_row = rrmap[rate]
                    else:
                        row += 1
                        res_row = rrmap[rate] = row
                        ws.cell(res_row, 2, rate)
                ws.cell(res_row, res_col,actuals[date][rate])
                ws.cell(res_row, 1, res_name)
        row += 1


def forecast_sheet(ws, resources, date_col):
    row = 2
    start_col = 3
    for res_name in resources.keys():
        ws.cell(row, 1, res_name)
        ws.cell(row,2, resources[res_name].getLastRate())
        forecast= res[res_name].getForecast()
        for date in forecast.keys():
            if date not in date_col.keys():
                date = getSunday(date)
            res_col = date_col[date] + start_col
            ws.cell(row, res_col,forecast[date])
        row += 1

if __name__ == '__main__':
    if len(sys.argv) >= 3:
        fcst = sys.argv[1]
        act = sys.argv[2]
        if len(sys.argv) >= 4:
            out = sys.argv[3]
        else:
            out = "TEST.xlsx"
        if fcst.endswith('.xlsx') and act.endswith('.xlsx'):
            try:
                os.stat(fcst)
                os.stat(act)
                res = dict()
                start_date = dt.date.today()
                end_date = start_date
                date_range = [start_date, end_date]
                proc_actuals(file=act, resources=res, se_dates=date_range)
                proc_forecast(file=fcst, resources=res, se_dates=date_range)
                start_date = date_range[0]
                end_date = date_range[1]
                if len(res) > 0:
                    wb = xl.Workbook()
                    dcmap = create_date_seq(start_date, end_date)
                    create_headers(wb.active, 'Actual', dcmap)
                    create_headers(wb.create_sheet(), 'Forecast', dcmap)

                    ws = wb['Actual']
                    actuals_sheet(ws, res, dcmap)
                    setColorToZeros(ws, PatternFill(patternType=None, start_color="FCD5B4", end_color="FCD5B4",
                                                    fill_type="solid"))
                    addFormulas(ws,max(dcmap.values()))
                    ws = wb['Forecast']
                    forecast_sheet(ws, res, dcmap)
                    addFormulas(ws,max(dcmap.values()))
                    wb.save(out)
                    wb.close()

                print('Done!')
            except WindowsError:
                print("Error: file does not exists")
        else:
            print("Error: files are not in OOXML formal .xlsx")
    else:
        print("Error: Invocation should be:\n python JDAProjActuals.py <forecast workbook path> <actuals workbook path>")




'''
Client Reporting Unit	| Client Name	| Project Name	| Client PO#	| Project Manager	| Associate Name	| AssociateId	| AssociateType	| Period Start Date	| Entry Date	| Task Name	| Total Hours	| Billable Hours	| Billing Rule Name	| Billable Amt. In Project Currency	| Project Currency Name	| Billable Amt. In USD	| Timesheet is Submitted	| Timesheet is Approved	| Timesheet Workflow State	| JTRAX Invoiced Status/Ref #	| Invoice Through Date	| Service Delivery Location	| Entry Note	| On Hold for Billing	| On Hold Reason

EL Rosado - Actuals - 02-Oct-2018.xlsx
EL Rosado - Forecast - 50216 - 02-Oct-2018.xlsx
2016-1114 WMS WLM Implem OP-0177020

"C:\\dev\\projects\\JDA\\internal\\PythonWorkshop\\data\\50216 - 2018 Octubre 1st v2.xlsx"
"C:\\dev\\projects\\JDA\\internal\\PythonWorkshop\\data\\CUST_JDA_MSRS_PS_PREINVOICE_TIME_SELPROJECT 2018-Oct-01.xlsx"
2015-1105 Unisuper PMM OP-0164785


'''


