import time
from datetime import date, timedelta, datetime
import pandas as pd
import openpyxl as xl
import xlrd
import os
from openpyxl.styles import PatternFill, Alignment, Font
import openpyxl.utils as xlu
import argparse

DF_ANAME = 'Associate Name'
DF_DATE  = 'Date'
DF_HOURS = 'Hours'
DF_COST  = 'Cost'
DF_RATE  = 'Rate'
DF_APPRV = 'Timesheet is Approved'

WS_THOURS = 'Total Hours'
WS_TCOST  = 'Total Cost'
WS_R0HRS = 'Total Hours Rate USD 0'
WS_R0ALT = 'Alt Total for Rate USD 0'
WS_TFCSTH = 'Total Forecast Hours'
WS_TFCSTC = 'Total Forecast Cost'
WS_RATE_RES = 'Resource'
WS_RATE_ACT_RATE = 'Actual Billing Rate'
WSN_ACTUALS = 'Actuals'
WSN_FCST = 'Forecast'
WSN_ACTFCST = 'Actuals+Forecast'
WSN_NOTAPPRV = 'Not Approved'
WSN_RATES = 'Rates'

ACTUALS2DROP= ['Client Reporting Unit', 'Client Name', 'Project Name', 'Client PO#', 'Project Manager',
               'AssociateId', 'AssociateType', 'Period Start Date', 'Task Name', 'Billable Hours', 'Billing Rule Name',
               'Billable Amt. In Project Currency', 'Project Currency Name', 'Timesheet is Submitted', 'Timesheet Workflow State',
               'JTRAX Invoiced Status/Ref #', 'Invoice Through Date', 'Service Delivery Location', 'Entry Note',
               'On Hold for Billing', 'On Hold Reason']
FCST2DROP = ['Role', 'Project', 'Actual Hours','User Last Name', 'User First Name']
#RATES2DROP = ['Resource Type','Hard Booked Hours','Forecasted Cost Rate','Forecasted Billing Rate','Actual Cost Rate']
RATES2DROP = ['Hard Booked Hours','Forecasted Cost Rate','Forecasted Billing Rate','Actual Cost Rate']

FORECAST_COLOR="CCFFCC"
TODAY_COLOR="00B050"
ZERO_COST_COLOR="FCD5B4"
DATES_COLOR="DCE6F1"
UNAPPRV_COLOR = "FF3300"

ZERO_FILLER = PatternFill(patternType=None, start_color=ZERO_COST_COLOR, end_color=ZERO_COST_COLOR,fill_type="solid")
FCST_FILLER = PatternFill(patternType=None, start_color=FORECAST_COLOR, end_color=FORECAST_COLOR,fill_type="solid")
TODAY_FILLER = PatternFill(patternType=None, start_color=TODAY_COLOR, end_color=TODAY_COLOR, fill_type="solid")
DATES_FILLER = PatternFill(patternType=None, start_color=DATES_COLOR, end_color=DATES_COLOR, fill_type="solid")
UNAPPRV_FILLER = PatternFill(patternType=None, start_color=UNAPPRV_COLOR, end_color=UNAPPRV_COLOR, fill_type="solid")

RATE_FORMULA = '=VLOOKUP(A{0},Rates!$A$2:$B${1},2,FALSE)'

'''
Returns the Sunday before pdate if pdate is not already Sunday
'''
def get_sunday(pdate):
    ndate = pdate
    #Start converting parameter to datetime type
    #   if it's a string it tries to interpret it as formatted as:
    #       mm/dd/yyyy
    #   if it's a date type it just crete a date time using parameter year, month and day
    if type(pdate) is str :
        if len(pdate) == 10:
            ndate = datetime(int(pdate[-4:]), int(pdate[0:2]), int(pdate[-7:-5]))
        else:
            ndate = None
    elif type(ndate) is date:
        ndate = datetime(ndate.year, ndate.month, ndate.day)

    #Get day of week, ie one of Monday to Sunday where Monday is 0,
    #Tuesday 1, Wenesday 2, Thursday 3,  Friday 4, Saturday 5 and Sunday 6
    dow = ndate.weekday()

    #If day of week is Sunday(6) we are OK
    #If it's not then we calculate a the date for previous Sunday
    if dow < 6:
        ndate = ndate - timedelta(dow+1)
    return ndate


''' 
Build headers for the worksheet. Has a row for resource name, if it contains actuals a column
with resource rate, one column for each week from the first week in project up to the last week in
the forecast. There are two columns after weeks one to sum hours and one to sum cost
'''
def create_headers(ws, title, dcmap, is_actual, max_name_len, is_faa):
    #Set worksheet name
    ws.title = title
    ws.insert_rows(1, 2)
    curr_col = 1
    font = Font(bold=True)
    ws.cell(1,curr_col,'Resources')
    if is_actual:
        curr_col += 1
        ws.cell(1, curr_col, DF_RATE)
    curr_col += 1
    r = Alignment(text_rotation=90, horizontal='center')
    for curr_date in dcmap.keys():
        c = ws.cell(1,curr_col,curr_date.date())
        c.fill = DATES_FILLER
        c.alignment = r
        c.font = font
        curr_col += 1
    c = ws.cell(1, curr_col, WS_THOURS)
    c.fill = DATES_FILLER
    r = Alignment(wrap_text=True, horizontal='center')
    c.alignment = r
    c.font = font
    ws.column_dimensions[c.column_letter].width = 13
    if is_actual:
        curr_col += 1
        c = ws.cell(1, curr_col, WS_TCOST)
        c.fill = DATES_FILLER
        c.alignment = r
        c.font = font
        ws.column_dimensions[c.column_letter].width = 13
        ws.freeze_panes = 'C2'
    else:
        ws.freeze_panes = 'B2'
    ws.column_dimensions['A'].width = max_name_len
    if is_faa:
        curr_col += 1
        c = ws.cell(1, curr_col, WS_R0HRS)
        c.fill = DATES_FILLER
        c.alignment = r
        c.font = font
        ws.column_dimensions[c.column_letter].width = 13
        curr_col += 1
        c = ws.cell(1, curr_col, WS_R0ALT)
        c.fill = DATES_FILLER
        c.alignment = r
        c.font = font
        ws.column_dimensions[c.column_letter].width = 13
        curr_col += 1
        c = ws.cell(1, curr_col, WS_TFCSTH)
        c.fill = DATES_FILLER
        c.alignment = r
        c.font = font
        ws.column_dimensions[c.column_letter].width = 13
        curr_col += 1
        c = ws.cell(1, curr_col, WS_TFCSTC)
        c.fill = DATES_FILLER
        c.alignment = r
        c.font = font
        ws.column_dimensions[c.column_letter].width = 13


"""
Processing for actuals sheet loaded in actuals DataFrame dat will be grouped by Associate Name
and rate. Rate 0 is set to a different color
"""
def actuals_sheet(ws, p_actuals, date_col, apprv):
    row = 1
    start_col = 3
    #Data is grouped by Consultant name, rate, and date all other data is summarized
    actuals_gb = p_actuals.groupby(by=[DF_ANAME, DF_RATE, DF_DATE]).sum()
    name = None
    rate = -1.0
    for index, group in actuals_gb.iterrows():
        if name != index[0] or rate != index[1]:
            row += 1
            ws.cell(row, 1, index[0])  # name
            ws.cell(row, 2, index[1]) #rate
            if index[1] == 0:
                set_color(ws.iter_cols(min_row=row, max_row=row, min_col=2),ZERO_FILLER)
        #Set a different color if time has not been approved
        if apprv and group[DF_APPRV] < 0:
            set_color(ws.iter_cols(min_row=row, max_row=row, min_col=2), UNAPPRV_FILLER)
        ws.cell(row,start_col + date_col[index[2]], group[DF_HOURS])
        #!!!!Not sure if next if is required. It lokks like the than the 2 lines above
        if apprv and group[DF_APPRV] < 0:
            set_color(ws.iter_cols(min_row=row, max_row=row, min_col=2), UNAPPRV_FILLER)
        name = index[0]
        rate = index[1]


"""
Processings for forecast sheet loaded in fcst DataFrame
"""
def forecast_sheet(ws, p_fcst, date_col):
    row = 1
    start_col = 2
    fcst_gb = p_fcst.groupby(by=[DF_ANAME, DF_DATE]).sum()
    curr_name = ''
    for index, group in fcst_gb.iterrows():
        name = index[0]
        if name != curr_name:
            row += 1
            ws.cell(row, 1, name)  # name
        ws.cell(row,start_col + date_col[index[1]], group[DF_HOURS])
        curr_name = name


def add_forecast(ws, row, name, p_fcst, rate_rows, fcst_date, start_col, date_col):
    if name is not None:
        ws.cell(row, 1, name)
        if rate_rows is not None:
            c = ws.cell(row, 2, RATE_FORMULA.format(row, rate_rows))
            c.number_format = '#,##0.00'
        set_color(ws.iter_cols(min_row=row, max_row=row, min_col=1), FCST_FILLER)
        res_fcst = p_fcst.loc[(p_fcst[DF_ANAME] == name) & (p_fcst[DF_DATE] > fcst_date)]
        res_fcst_gb = res_fcst.groupby(by=[DF_ANAME, DF_DATE]).sum()
        for index_fcst, group_fcst in res_fcst_gb.iterrows():
            ws.cell(row, start_col + date_col[index_fcst[1]], group_fcst[DF_HOURS])


"""
Combining Actuals and Forecast data in one sheet. Actuals data fills the sheet up to last week. Since
last week to last week with forecast data will fill next weeks. Rate for forecast has to be set by 
spradsheet user. Special colors ha been set for rate value 0 and for forecast rows.
"""
def fcst_act_sheet(ws, p_fcst, p_actuals, date_col, rate_rows):
    """Forecast date is set as the last week having actuals even if it should be next week"""
    fcst_date = p_actuals[DF_DATE].max()
    row = 1
    start_col = 3
    """Acual data is grouped by resource, rate and week summing up reported time"""
    actuals_gb = p_actuals.groupby(by=[DF_ANAME, DF_RATE, DF_DATE]).sum()
    name = None
    rate = -1.0
    for index, group in actuals_gb.iterrows():
        """When there's a change of resource or rate we add a new row"""
        if name != index[0] or rate != index[1]:
            row += 1
            if name != index[0]:
                add_forecast(ws, row, name, p_fcst, rate_rows, fcst_date, start_col, date_col)
                if name is not None:
                    row += 1
                name = index[0]
            rate = index[1]
            ws.cell(row, 1, name)
            ws.cell(row, 2, rate)
            if rate == 0:
                set_color(ws.iter_cols(min_row=row, max_row=row, min_col=2),ZERO_FILLER)
        ws.cell(row,start_col + date_col[index[2]], group[DF_HOURS])
        if group[DF_APPRV] < 0:
            set_color(ws.iter_cols(min_row=row, max_row=row, min_col=2), UNAPPRV_FILLER)

    if name is not None:
        row += 1
        add_forecast(ws, row, name, p_fcst, rate_rows, fcst_date, start_col, date_col)

    res_fcst = p_fcst.loc[p_fcst[DF_DATE] >= fcst_date]
    res_fcst_gb=res_fcst[~res_fcst[DF_ANAME].isin(p_actuals[DF_ANAME])].groupby(by=[DF_ANAME, DF_DATE]).sum()
    name = None
    for index_fcst, group_fcst in res_fcst_gb.iterrows():
        if name != index_fcst[0]:
            name = index_fcst[0]
            row += 1
            ws.cell(row, 1, name)
            set_color(ws.iter_cols(min_row=row, max_row=row, min_col=1), FCST_FILLER)

        ws.cell(row, start_col + date_col[index_fcst[1]], group_fcst[DF_HOURS])

    if fcst_date in date_col:
        col = date_col[fcst_date] + 3
        set_color(ws.iter_cols(min_row=2, min_col=col, max_col=col),TODAY_FILLER)


"""
Returns a weekly date sequence
"""
def create_date_seq(p_start, p_end):
    seq = dict()
    shift = 0
    for v_date in pd.date_range(start=p_start, end=p_end, freq='W').to_pydatetime():
        seq[v_date] = shift
        shift += 1
    return seq


"""
Set a filller color for a set of cells
"""
def set_color(p_iter, fill):
    for sset in p_iter:
        for cell in sset:
            cell.fill = fill


"""
Parse rate data
"""
def parse_rate(s):
    if s.strip() == 'None':
        return None
    else:
        try:
            rate=float(s.strip('USD $').strip('/Hour'))
            return rate
        except ValueError:
            return 0.0


"""
Add sum formula to all rows with data
"""
def add_formulas(ws, shift, is_actual, is_faa):
    start_col = 3 if is_actual else 2
    last_col = ws.cell(1, start_col+shift).column
    hours_formula = "=SUM({0}{1}:{2}{1})"
    autosum_formula = "=SUM({0}2:{0}{1})"
    cost_formula = "={0}{1}*B{1}"
    hours = start_col + shift + 1
    cost = hours + 1
    for r in ws.iter_rows(min_row=2):
        if is_faa:
            c = ws.cell(r[0].row,start_col -1)
            #Totalizing rate 0 actauals
            if c.fill == ZERO_FILLER or (c.fill == UNAPPRV_FILLER and c.value == 0):
                c = ws.cell(r[0].row, hours + 2, hours_formula.format('C', r[0].row, xlu.cell.get_column_letter(last_col)))
                c.number_format = '#,##0.00'
                h_col = c.column
                c = ws.cell(r[0].row,cost + 2, cost_formula.format(xlu.cell.get_column_letter(h_col), r[0].row))
                c.number_format = '#,##0.00'
            #Totalizing forecast
            elif c.fill == FCST_FILLER:
                c = ws.cell(r[0].row, hours + 4, hours_formula.format('C', r[0].row, xlu.cell.get_column_letter(last_col)))
                c.number_format = '#,##0.00'
                h_col = c.column
                c = ws.cell(r[0].row,cost + 4, cost_formula.format(xlu.cell.get_column_letter(h_col), r[0].row))
                c.number_format = '#,##0.00'
            else:
                c = ws.cell(r[0].row, hours, hours_formula.format('C', r[0].row, xlu.cell.get_column_letter(last_col)))
                c.number_format = '#,##0.00'
                if is_actual:
                    h_col = c.column
                    c = ws.cell(r[0].row,cost, cost_formula.format(xlu.cell.get_column_letter(h_col), r[0].row))
                    c.number_format = '#,##0.00'
        else:
            c = ws.cell(r[0].row, hours, hours_formula.format('C', r[0].row, xlu.cell.get_column_letter(last_col)))
            c.number_format = '#,##0.00'
            if is_actual:
                c = ws.cell(r[0].row, cost, cost_formula.format(xlu.cell.get_column_letter(hours), r[0].row))
                c.number_format = '#,##0.00'

    c = ws.cell(ws.max_row+1, hours, autosum_formula.format(xlu.cell.get_column_letter(hours), ws.max_row)) #c.row-1))
    c.number_format = '#,##0.00'
    font = Font(bold=True)
    c.font = font
    if is_actual:
        #c = ws.cell(ws.max_row, cost, autosum_formula.format(xlu.cell.get_column_letter(c.column), c.row-1))
        c = ws.cell(ws.max_row, cost, autosum_formula.format(xlu.cell.get_column_letter(cost), ws.max_row - 1))
        c.number_format = '#,##0.00'
        c.font = font
        if is_faa:
            col = cost +1
            #c = ws.cell(ws.max_row,col, autosum_formula.format(xlu.cell.get_column_letter(c.column), c.row - 1))
            c = ws.cell(ws.max_row, col, autosum_formula.format(xlu.cell.get_column_letter(col), ws.max_row - 1))
            c.number_format = '#,##0.00'
            c.font = font
            col += 1
            #c = ws.cell(ws.max_row,col, autosum_formula.format(xlu.cell.get_column_letter(c.column), c.row - 1))
            c = ws.cell(ws.max_row, col, autosum_formula.format(xlu.cell.get_column_letter(col), ws.max_row - 1))
            c.number_format = '#,##0.00'
            c.font = font
            col += 1
            #c = ws.cell(ws.max_row,col, autosum_formula.format(xlu.cell.get_column_letter(c.column), c.row - 1))
            c = ws.cell(ws.max_row, col, autosum_formula.format(xlu.cell.get_column_letter(col), ws.max_row - 1))
            c.number_format = '#,##0.00'
            c.font = font
            col += 1
            #c = ws.cell(ws.max_row,col, autosum_formula.format(xlu.cell.get_column_letter(c.column), c.row - 1))
            c = ws.cell(ws.max_row, col, autosum_formula.format(xlu.cell.get_column_letter(col), ws.max_row - 1))
            c.number_format = '#,##0.00'
            c.font = font



def process(p_fcst, p_actuals, p_rates):
    '''
    Forecast columns are:
    .__________________________________.
    |   Column data      |Column index |
    .__________________________________.
    |User Last Name      |     0       |
    |User First Name     |     1       |
    |Role                |     2       |
    |Project             |     3       |
    |Date                |     4       |
    |Actual Hours        |     5       |
    |Total Booking Hours |     6       |
    .__________________________________.

    Relevant data to be kept is consultant name (columns 0 and 1), date (columns 4) and booking hours (column 6).
    Data is consolidated by week. Weeks start on Sunday.
    First row with data is 9.
    '''

    # A new column is created as the concatenation of last and first names
    p_fcst[DF_ANAME] = p_fcst['User Last Name'] + ", " + p_fcst['User First Name']
    # When there is no resource assign (ie: last name is NA) the Associate Name column is filled with the Role description
    p_fcst[DF_ANAME].fillna(p_fcst['Role'], inplace=True)
    # Housekeeping, all unnecessary columns are dropped
    p_fcst.drop(FCST2DROP, 1)
    # Renaming column
    p_fcst = p_fcst.rename(index=str, columns={"Total Booking Hours": DF_HOURS})

    from_date = p_fcst[DF_DATE].min().date()
    to_date = p_fcst[DF_DATE].max().date()
    max_name_len = p_fcst[DF_ANAME].map(len).max()

    '''
    Actuals columns are:
    .__________________________________________________.
    |         Column data               | Column index |
    .__________________________________________________.
    | Client Reporting Unit            	|      0       |
    | Client Name	                    |      1       | 
    | Project Name	                    |      2       |
    | Client PO#	                    |      3       |
    | Project Manager	                |      4       |
    | Associate Name	                |      5       |
    | AssociateId	                    |      6       |
    | AssociateType	                    |      7       |
    | Period Start Date	                |      8       |
    | Entry Date	                    |      9       |
    | Task Name	                        |     10       |
    | Total Hours	                    |     11       |
    | Billable Hours	                |     12       |
    | Billing Rule Name	                |     13       |
    | Billable Amt. In Project Currency	|     14       |
    | Project Currency Name	            |     15       |
    | Billable Amt. In USD	            |     16       |
    | Timesheet is Submitted	        |     17       |
    | Timesheet is Approved           	|     18       |
    | Timesheet Workflow State	        |     19       |
    | JTRAX Invoiced Status/Ref #	    |     20       |
    | Invoice Through Date	            |     21       |
    | Service Delivery Location	        |     22       |
    | Entry Note	                    |     23       |
    | On Hold for Billing	            |     24       |
    | On Hold Reason                    |     25       |
    .__________________________________________________.
    Relevant data to be kept is associate name (column 5), entry date (column 9),  hours (column 11), amount in USD (column 16).
    Amount by itself is not kept, it is used to calculate rate (rate= <amount> / <hours>). Hours are consolidated by week/rate.
    Weeks start on Sunday.
    First row with data is 9. 
    '''
    p_actuals.drop(ACTUALS2DROP, 1)
    p_actuals = p_actuals.rename(index=str,
                             columns={"Entry Date": DF_DATE, "Total Hours": DF_HOURS, "Billable Amt. In USD": DF_COST})
    dt = p_actuals[DF_DATE].min().date()
    if dt < from_date:
        from_date = dt
    dt = p_actuals[DF_DATE].max().date()
    if dt > to_date:
        to_date = dt
    tmp_len = p_actuals[DF_ANAME].map(len).max()
    if tmp_len > max_name_len:
        max_name_len = tmp_len
    p_actuals[DF_RATE] = p_actuals[DF_COST] / p_actuals[DF_HOURS]

    v_wb = xl.Workbook()
    dcmap = create_date_seq(from_date, to_date)
    create_headers(v_wb.active, WSN_ACTUALS, dcmap, True, max_name_len,False)
    create_headers(v_wb.create_sheet(), WSN_FCST, dcmap, False, max_name_len,False)
    create_headers(v_wb.create_sheet(), WSN_ACTFCST, dcmap, True, max_name_len,True)

    '''
    Rates columns are:
    .__________________________________.
    |   Column data         |Column index |
    ._____________________________________.
    |Resource               |     0       |
    |Resource Type          |     1
    |Hard Booked Hours      |     2       |
    |Forecasted Cost Rate   |     3       |
    |Forecasted Billing Rate|     4       |
    |Actual Cost Rate       |     5       |
    |Actual Billing Rate    |     6       |
    ._____________________________________.

    Relevant data to be kept is consultant name (column 0), actual billing rate (columns 5).
    First row with data is 1.
    '''
    r_rows = None
    if p_rates is not None:
        p_rates = p_rates.drop(RATES2DROP,1)
        tmp_len = p_rates['Resource'].map(len).max()
        if tmp_len > max_name_len:
            max_name_len = tmp_len
        #Rates spreadsheet
        ratesws = v_wb.create_sheet()
        ratesws.title = WSN_RATES
        row = 1
        c = ratesws.cell(row, 1, WS_RATE_RES)
        c.alignment = Alignment(wrap_text=True, horizontal='center')
        c.font = Font(bold=True)
        c = ratesws.cell(row, 2, WS_RATE_ACT_RATE)
        c.alignment = Alignment(wrap_text=True, horizontal='center')
        c.font = Font(bold=True)

        ratesws.column_dimensions['A'].width = max_name_len

        row += 1
        for index, data in p_rates.iterrows():
            ratesws.cell(row, 1,data[WS_RATE_RES])
            ratesws.cell(row, 2, data[WS_RATE_ACT_RATE])
            row +=1
        r_rows = row - 1

    actuals_sheet(v_wb[WSN_ACTUALS], p_actuals, dcmap, True)
    add_formulas(v_wb[WSN_ACTUALS], max(dcmap.values()), True, False)

    forecast_sheet(v_wb[WSN_FCST], p_fcst, dcmap)
    add_formulas(v_wb[WSN_FCST], max(dcmap.values()), False, False)

    fcst_act_sheet(v_wb[WSN_ACTFCST], p_fcst, p_actuals, dcmap, r_rows)
    add_formulas(v_wb[WSN_ACTFCST], max(dcmap.values()), True, True)

    df_napprv = p_actuals[p_actuals[DF_APPRV] < 0]
    if df_napprv.size > 0:
        from_date = df_napprv[DF_DATE].min().date()
        to_date = df_napprv[DF_DATE].max().date()
        dcmap = create_date_seq(from_date, to_date)
        create_headers(v_wb.create_sheet(), WSN_NOTAPPRV, dcmap, True, max_name_len,False)
        actuals_sheet(v_wb[WSN_NOTAPPRV], df_napprv, dcmap, False)
        add_formulas(v_wb[WSN_NOTAPPRV], max(dcmap.values()), True, False)
    return v_wb

if __name__ == '__main__':
    start = time.time()
    parser = argparse.ArgumentParser(description='JDA PM tool for actuals + forecast')
    parser.add_argument('--fcst', '-fcst', required=True, help='Forecast data spreadsheet (.xls or .xlsx file)')
    parser.add_argument('--act', '-act', required=True, help='Actuals data spreadsheet (.xls or .xlsx file)')
    parser.add_argument('--out', '-out', help='Output Excel file (.xlsx)', default='JDAFcstActuals.xlsx')
    parser.add_argument('--rates', '-rates', help='Rates data spreadsheet (.xls or .xlsx file)')
    params = parser.parse_args()

    out = params.out
    fcstxl = params.fcst
    actxl = params.act
    ratesxl = params.rates
    if (fcstxl.endswith('.xlsx') or fcstxl.endswith('.xls')) and (actxl.endswith('.xlsx') or actxl.endswith('.xls')):
        try:
            os.stat(fcstxl)
        except WindowsError:
            print("ERROR: file \n\t{0} \ndoes not exists\n\n".format(fcstxl))
            input('Press ENTER to exit')
            exit(-1)
        try:
            os.stat(actxl)
        except WindowsError:
            print("ERROR: file \n{0} \ndoes not exists\n\n".format(actxl))
            input('Press ENTER to exit')
            exit(-11)
        try:
            if fcstxl.endswith('.xls'):
                book = xlrd.open_workbook(fcstxl)
                fcst = pd.read_excel(book, engine='xlrd', header=6, converters={4: get_sunday})
            else:
                fcst = pd.read_excel(fcstxl, header=6, converters={4: get_sunday})
        except Exception:
            print("There's something wrong with file format for \n{0}".format(fcstxl))
            input('Press ENTER to exit')
            exit(-2)
        try:
            if actxl.endswith('.xls'):
                book = xlrd.open_workbook(actxl)
                actuals = pd.read_excel(book, engine='xlrd', header=6, converters={'Entry Date': get_sunday,'Timesheet is Approved':lambda x: 0 if x == 'Y' else -1})
            else:
                actuals = pd.read_excel(actxl,header=7, converters={'Entry Date': get_sunday,'Timesheet is Approved':lambda x: 0 if x == 'Y' else -1})
        except Exception:
            print("There's somthing wrong with file format for \n{0}".format(actxl))
            input('Press ENTER to exit')
            exit(-2)
        try:
            if ratesxl is None:
                ratesxl = None
            else:
                if ratesxl.endswith('.xls'):
                    book = xlrd.open_workbook(ratesxl)
                    rates = pd.read_excel(book, engine='xlrd', header=0, converters={'Actual Billing Rate': parse_rate})
                else:
                    rates = pd.read_excel(ratesxl,header=0, converters={'Actual Billing Rate': parse_rate})
        except Exception:
            print("There's somthing wrong with file format for \n{0}".format(actxl))
            input('Press ENTER to exit')
            exit(-2)
        #try:
        wb = process(fcst,actuals,rates)
        wb.save(out)
        wb.close()
        print('Done!')
        #except Exception as excp:
        #    print(excp)
        #    input('Press ENTER to exit')
        #    exit(-2)
    else:
        print("Error: files are not in OOXML format .xlsx nor BIFF format .xls")
    print("Executed in {0:.5f} seconds".format((time.time()-start)))
